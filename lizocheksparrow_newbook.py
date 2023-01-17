# This Python 3 environment comes with many helpful analytics libraries installed

# It is defined by the kaggle/python Docker image: https://github.com/kaggle/docker-python

# For example, here's several helpful packages to load



import numpy as np # linear algebra

import pandas as pd # data processing, CSV file I/O (e.g. pd.read_csv)



# Input data files are available in the read-only "../input/" directory

# For example, running this (by clicking run or pressing Shift+Enter) will list all files under the input directory



import os

for dirname, _, filenames in os.walk('/kaggle/input'):

    for filename in filenames:

        print(os.path.join(dirname, filename))



# You can write up to 5GB to the current directory (/kaggle/working/) that gets preserved as output when you create a version using "Save & Run All" 

# You can also write temporary files to /kaggle/temp/, but they won't be saved outside of the current sessionзалогинился
import re
from openpyxl import load_workbook

from openpyxl.utils.exceptions import InvalidFileException

from os import listdir

from time import localtime, mktime, strptime

from numpy import array, dtype

datadir = '/kaggle/input/lv-200602/'
listdir('/kaggle/input/lv-200602/')
from os import getcwd

getcwd()


from sklearn.model_selection import train_test_split

from sklearn.neural_network import MLPRegressor

# from sklearn.preprocessing import Normalizer
holidays = {2018: {

    1: [1, 2, 3, 4, 5, 8],

    2: [23],

    3: [8, 9],

    5: [1, 2, 9],

    6: [11, 12],

    11: [5],

    12: [31]},

    2017: {

        1: [2, 3, 4, 5, 6],

          2: [23, 24],

3: [8],

5: [1, 8, 9],

6: [12],

11: [6]},

    2019: {

    1: [1, 2, 3, 4, 7, 8],

    3: [8],

    5: [1, 2, 3, 9, 10],

    6: [12],

    11: [4]},

    2020: {

        6: [12],

        11: [4]

    }

}

#  рабочие дни в week-end

unholidays = {

    2018: {4: [28], 6: [9], 12: [29]},

    2017: {}

}
#%%

days_nums_in_months = {

    1: 31,

    2: 28,

    3: 31,

    4: 30,

    5: 31,

    6: 30,

    7: 31,

    8: 31,

    9: 30,

    10: 31,

    11: 30,

    12: 31

}
def get_list_normalized(list_):

    min_ = min(list_)

    range_ = max(list_) - min_

    return [(el - min_) / range_ for el in list_]
def get_matrix_normalized(matrix):

    new_matrix = matrix.copy()

    new_matrix.dtype = dtype('float64')

    rows_num, cols_num = matrix.shape

    i_range = range(rows_num)

    for j in range(cols_num):

        list2normalize = [matrix[i, j] for i in i_range]

        normalized_list = get_list_normalized(list2normalize)

        for i in i_range:

            new_matrix[i, j] = normalized_list[i]

    return new_matrix
def denormalize(value, min_, range_):

    return min_ + value * range_
def get_timestamp(start_struct, stop_struct):

    start_stamp = mktime(start_struct)

    stop_stamp = mktime(stop_struct)

    return (start_stamp + stop_stamp) / 2
def timestamp(f):

    def ret(start_struct, stop_struct):

        timestamp = get_timestamp(start_struct, stop_struct)

        return f(timestamp)

    return ret
#%%

@timestamp

def get_midmonth(timestamp):

    day = localtime(timestamp)[2]

    return day in range(11, 25)
@timestamp

def get_day_in_year(timestamp):

    return localtime(timestamp)[-2]
@timestamp

def reciprocal_day_in_year(timestamp):

    return 365 - localtime(timestamp)[-2]
@timestamp

def get_winter(timestamp):

    return localtime(timestamp)[1] in (1, 2, 12)
@timestamp

def get_spring(timestamp):

    return localtime(timestamp)[1] in set([i for i in range(3, 6)])
@timestamp

def get_summer(timestamp):

    return localtime(timestamp)[1] in set([i for i in range(6, 9)])
@timestamp

def get_fall(timestamp):

    return localtime(timestamp)[1] in set([i for i in range(9, 12)])

#%%
def get_days_num(start, stop):

    diff = stop.tm_yday - start.tm_yday + 1

    if diff < 0:

        return diff + 365

    else:

        return diff

#%%
def get_hol_num(start, stop):

    hol_num = 0

    year = start.tm_year

    month = start.tm_mon

    day = start.tm_mday

    week_day = start.tm_wday

    to_iterate = True

    while to_iterate:

        try:

            list_ = holidays[year][month]

            if day in list_:

                hol_num += 1

        except KeyError:

            pass

        try:

            list_ = unholidays[year][month]

            if day in list_:

                hol_num -= 1

        except KeyError:

            pass

        if week_day in (5, 6):

            hol_num += 1



        week_day = (week_day + 1) % 7

        if day == stop.tm_mday:

            to_iterate = False

        day += 1

        if day > days_nums_in_months[month]:

            month += 1

            day = 1

        if month > 12:

            year += 1

            month = 1

    return hol_num
def get_hol_part(start, stop):

    return get_hol_num(start, stop) / get_days_num(start, stop)
@timestamp

def day_in_month(timestamp):

    return localtime(timestamp).tm_mday
@timestamp

def part_of_month(timestamp):

    struct_stamp = localtime(timestamp)

    return struct_stamp.tm_mday / days_nums_in_months[struct_stamp.tm_mon]
@timestamp

def reciprocal_part_of_month(timestamp):

    struct_stamp = localtime(timestamp)

    return 1 - struct_stamp.tm_mday / days_nums_in_months[struct_stamp.tm_mon]
@timestamp

def day_in_week(timestamp):

    return localtime(timestamp).tm_wday
@timestamp

def reciprocal_day_in_week(timestamp):

    return 7 - localtime(timestamp).tm_wday
#%%

features_to_use = [

    get_midmonth,

    get_day_in_year,

    # get_winter,

    # get_spring,

    # get_summer,

    # get_fall,

    get_hol_part,

    # day_in_month,

    day_in_week,

    part_of_month,

    reciprocal_part_of_month,

    reciprocal_day_in_week,

    reciprocal_day_in_year

]
#%%

rex = re.compile(r'.*(\d\d)\.(\d\d)\.(\d{4}) - (\d\d)\.(\d\d)\.(\d{4})')

start_range = range(1, 4)

stop_range = range(4, 7)
def get_current_x_and_days_num(string):

    match_obj = rex.match(string)

    if match_obj:

        start_tuple = tuple([int(match_obj.group(i)) for i in start_range])

        start_struct = strptime('%s.%s.%s' % start_tuple, '%d.%m.%Y')

        stop_tuple = tuple([int(match_obj.group(i)) for i in stop_range])

        stop_struct = strptime('%s.%s.%s' % stop_tuple, '%d.%m.%Y')

        days_num = get_days_num(start_struct, stop_struct)

        return [feature(start_struct, stop_struct) for feature in

                features_to_use], days_num
#%%

#  загружаем данные

atms = []

file_names = listdir(datadir)

for file_name in file_names:

    ds2append = [], []

    path = datadir + file_name

    try:

        wb = load_workbook(filename=path)

        sheet_name = wb.sheetnames[0]

        sheet = wb[sheet_name]

        j = 2

        while sheet[f'C{j}'].value:

            value = sheet[f'C{j}'].value

            a = sheet[f'A{j}'].value

            b = sheet[f'B{j}'].value

            if value.startswith('Выдача'):

                current_x, days_num = get_current_x_and_days_num(value)

                current_y = int(b) / days_num

                ds2append[0].append(current_x)

                ds2append[1].append(current_y)

            j += 1

        atms.append({'x': array(ds2append[0]), 'y': ds2append[1]})

    except InvalidFileException:

        pass
for atm in atms:

    shape = atm['x'].shape

    i_range, j_range = [range(shape[i]) for i in range(2)]

    atm['min'] = min(atm['y'])

    atm['range'] = max(atm['y']) - atm['min']

    atm['mins'] = [min([atm['x'][i, j] for i in i_range]) for j in j_range]

    atm['ranges'] = [max([atm['x'][i, j] for i in i_range]) - atm['mins'][j]

                     for j in j_range]

    atm['x'] = get_matrix_normalized(atm['x'])  # нормализуем

    atm['y'] = get_list_normalized(atm['y'])    # данные
#%%

atms[0]['layers'] = 34, 14, 21

# atms[0]['layers'] = 8, 50, 1

atms[1]['layers'] = 16, 11

atms[2]['layers'] = 10, 5, 18, 17

atms[3]['layers'] = 32, 50, 38

atms[4]['layers'] = 20, 38

atms[5]['layers'] = 39, 29

atms[6]['layers'] = 15, 13
for i in range(len(file_names)):

    print(f'atm_{i} --- {file_names[i]}')
#%%

atm_ix = 0  # выбираем банкомат

atm = atms[atm_ix]

regressor = MLPRegressor(hidden_layer_sizes=atm['layers'],

                         random_state=1, activation='tanh')

regressor.fit(atm['x'], atm['y'])  # обучение
#%%

min_ = atm['min']

range_ = atm['range']

time_range_for_prediction = '31.05.2020 - 31.05.2020'  # период, на который

                                                       # надо спрогнозировать

                                                       # загрузку

x, days_num = get_current_x_and_days_num(time_range_for_prediction)

x = array([x])

new_x = x.copy()

new_x.dtype = dtype('float64')

for j in range(x.shape[1]):

    new_x[0, j] = (x[0, j] - atm['mins'][j]) / atm['ranges'][j]

y = regressor.predict(new_x)

prediction = denormalize(y, min_, range_) * days_num

print(f'Рекомендовано загрузить {prediction[0]}')