import pandas as pd
import numpy as np
from bs4 import BeautifulSoup
from urllib import request
from urllib.error import URLError, HTTPError
import re
import lxml.html
import requests
import urllib
import ssl
import urllib.parse
import itertools
import openpyxl

import pdfminer
import io 
import PyPDF2
from PyPDF2 import PdfFileReader
myswdurl = pd.read_excel("yourdata.xlsx",
                      sheet_name ='SWDlinks',usecols='a',
                      encoding='utf-8',skipinitialspace=True,sep=",")
mydocurl = pd.read_excel("yourdata.xlsx",
                      sheet_name ='DOClinks',usecols='a',
                      encoding='utf-8',skipinitialspace=True,sep=",")
mypdfurl = pd.read_excel("yourdata.xlsx",
                      sheet_name ='PDFlinks',usecols='a',
                      encoding='utf-8',skipinitialspace=True,sep=",")
regex = re.compile(
        r'^(?:http|ftp)s?://' # http:// or https://
        r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?)|' #domain...
        r'localhost|' #localhost...
        r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})' # ...or ip
        r'(?::\d+)?' # optional port
        r'(?:/?|[/?]\S+)$', re.IGNORECASE)

records = []
# for Sofware links

for index,row in myswdurl.itertuples():
    if re.match(regex,row) is not None:
        try:
            headers = {'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.94 Safari/537.36'}
            swdhtml_page = request.urlopen(row) # for opening url from excel
            
            _create_unverified_https_context = ssl._create_unverified_context # for ssl certificate issue
        #ignore errors
        except HTTPError: 
            pass
        except URLError:
            pass
        except requests.exceptions.ConnectionError as e:
            pass
        else:
            ssl._create_default_https_context = _create_unverified_https_context
            
        #web link parsing    
        soup = BeautifulSoup(swdhtml_page,'lxml') # parsing pages
        
        #assinging value to variables.
        mytitles = soup.find('div',{'class':'wrapper'})
                
        if mytitles is not None:
            for t in mytitles:
                try:
                    myjstitle = t.find('h2').text
                except AttributeError:
                    continue
                    
            records.append({'Link':row,'Web Titles':myjstitle})
            
            print(row,myjstitle)
                
        srcfile = openpyxl.load_workbook('yourfilename.xlsx',read_only=False)
        sheetname = srcfile.get_sheet_by_name('Sheet1')
                
        for index, descr in enumerate(records):
            rowNum = sheetname.max_row + 1
            sheetname.cell(row=rowNum,column = 1).value = str(row)
            sheetname.cell(row=rowNum,column = 2).value = myjstitle
                
        srcfile.save('yourfilename.xlsx') 
regex = re.compile(
        r'^(?:http|ftp)s?://' # http:// or https://
        r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?)|' #domain...
        r'localhost|' #localhost...
        r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})' # ...or ip
        r'(?::\d+)?' # optional port
        r'(?:/?|[/?]\S+)$', re.IGNORECASE)

records = []

#for Document links
                
for index,row in mydocurl.itertuples():
    if re.match(regex,row) is not None:
        try:
            headers = {'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.94 Safari/537.36'}
            dochtml_page = request.urlopen(row) # for opening url from excel
            
            _create_unverified_https_context = ssl._create_unverified_context # for ssl certificate issue
            #ignore errors
        except HTTPError: 
            pass
        except URLError:
            pass
        except requests.exceptions.ConnectionError as e:
            pass
        else:
            ssl._create_default_https_context = _create_unverified_https_context
            
            #web link parsing    
        soup = BeautifulSoup(dochtml_page,'lxml') # parsing pages
        
            #assinging value to variables.
        pdftitles = soup.find('b')
          
        if pdftitles is not None:
            for pdf in pdftitles:
                try:
                    mypdftitles = pdftitles.next.next
                except AttributeError:
                    continue

            records.append({'Link':row,'Web Titles':mypdftitles})
                
        print(row,mypdftitles)
        
        srcfile = openpyxl.load_workbook('yourfilename.xlsx',read_only=False)
        #sheetname = srcfile.sheetname['Sheet1']
        sheetname = srcfile.get_sheet_by_name('Sheet1')
                
        for index, descr in enumerate(records):
            rowNum = sheetname.max_row + 1
            sheetname.cell(row=rowNum,column = 1).value = str(row)
            sheetname.cell(row=rowNum,column = 2).value = mypdftitles
                                           
        srcfile.save('yourfilename.xlsx')              
                
regex = re.compile(
        r'^(?:http|ftp)s?://' # http:// or https://
        r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?)|' #domain...
        r'localhost|' #localhost...
        r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})' # ...or ip
        r'(?::\d+)?' # optional port
        r'(?:/?|[/?]\S+)$', re.IGNORECASE)

records = []

#for PDF links
                
for index,row in mypdfurl.itertuples():
    if re.match(regex,row) is not None:
        try:
            headers = {'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.94 Safari/537.36'}
            pdfhtml_page = request.urlopen(row) # for opening url from excel
            
            _create_unverified_https_context = ssl._create_unverified_context # for ssl certificate issue
            #ignore errors
        except HTTPError: 
            pass
        except URLError:
            pass
        except requests.exceptions.ConnectionError as e:
            pass
        else:
            ssl._create_default_https_context = _create_unverified_https_context
            
            #web link parsing
        r = requests.get(row)
                               
        f = io.BytesIO(r.content)
                   
        #reader = PdfFileReader(f)
        
        try:
            reader = PdfFileReader(f)
            contents = reader.documentInfo.title
        
        except ConnectionError:
            continue
        except ConnectionRefusedError:
            continue
        print(row,contents)
        
        records.append({'Links':row,'Web Titles':contents})
        srcfile = openpyxl.load_workbook('yourfilename.xlsx',read_only=False)
        sheetname = srcfile.get_sheet_by_name('Sheet1')
                                
        for index, descr in enumerate(records):
            rowNum = sheetname.max_row + 1
            sheetname.cell(row=rowNum,column = 1).value = str(row)
            sheetname.cell(row=rowNum,column = 2).value = contents
                
        srcfile.save('yourfilename.xlsx')        