#!pip install scikit-learn==0.22.2
!apt-get install build-essential swig
!curl https://raw.githubusercontent.com/automl/auto-sklearn/master/requirements.txt | xargs -n 1 -L 1 pip install
!pip install -U auto-sklearn
#!conda create -n autopytorch -y
#!pip install torchvision pytorch
!git clone https://github.com/automl/Auto-PyTorch.git
!cd Auto-PyTorch
!cat requirements.txt | xargs -n 1 -L 1 pip install
!python setup.py install
!git clone https://github.com/automl/Auto-PyTorch.git
!cd Auto-PyTorch
!cat Auto-PyTorch/requirements.txt | xargs -n 1 -L 1 pip install

!python /setup.py install
from autoPyTorch import AutoNetClassification
# This Python 3 environment comes with many helpful analytics libraries installed
# It is defined by the kaggle/python Docker image: https://github.com/kaggle/docker-python
# For example, here's several helpful packages to load

from sklearn.model_selection import train_test_split
import numpy as np # linear algebra
import pandas as pd # data processing, CSV file I/O (e.g. pd.read_csv)
from PIL import Image
from random import randrange
#from autoPyTorch import AutoNetClassification

import sklearn.model_selection
import sklearn.datasets
import sklearn.metrics
from sklearn.neural_network import MLPClassifier

import torchvision.models as models

from fastai import *
from fastai.vision import *
from fastai.metrics import error_rate

import cv2

# You can write up to 5GB to the current directory (/kaggle/working/) that gets preserved as output when you create a version using "Save & Run All" 
# You can also write temporary files to /kaggle/temp/, but they won't be saved outside of the current session
img = cv2.imread('../input/surfacetestimages/Risse_in_der_Wand.jpg')

kernel = np.ones((5,5),np.float32)/25
#dst = cv2.filter2D(img,-1,kernel)
blur = cv2.bilateralFilter(img,9,75,75)

plt.subplot(121),plt.imshow(img),plt.title('Original')
plt.xticks([]), plt.yticks([])
plt.subplot(122),plt.imshow(blur),plt.title('Averaging')
plt.xticks([]), plt.yticks([])
plt.show()
!pip install -U Pillow
from PIL import Image
# Input data files are available in the read-only "../input/" directory
# For example, running this (by clicking run or pressing Shift+Enter) will list all files under the input directory
data = []
limit = 10000
count = 0

import os
for dirname, _, filenames in os.walk('/kaggle/input/surface-crack-detection/Negative'):
    for filename in filenames:
        data.append(np.asarray(Image.open(os.path.join(dirname, filename)).convert(mode='L')))
        count += 1
        if count == limit:
            break
        #print(os.path.join(dirname, filename))

count = 0

for dirname, _, filenames in os.walk('/kaggle/input/surface-crack-detection/Positive'):
    for filename in filenames:
        data.append(np.asarray(Image.open(os.path.join(dirname, filename)).convert(mode='L')))
        count += 1
        if count == limit:
            break
        #print(os.path.join(dirname, filename))
        
dataArr = np.asarray(data).reshape(2*limit, 51529)
labels = np.asarray(np.zeros(limit).tolist()+np.ones(limit).tolist())

#blur = cv2.bilateralFilter(dataArr)

X_train, X_test, y_train, y_test = train_test_split(
        dataArr,
        labels,
        test_size=0.2,
        shuffle=True,
        random_state=randrange(2000))
import autosklearn.classification
automl = autosklearn.classification.AutoSklearnClassifier(ml_memory_limit=12288)
automl.fit(X_train, y_train)
y_hat = automl.predict(X_test)
print("Accuracy score ", sklearn.metrics.accuracy_score(y_test, y_hat))
print("F1 score: ", sklearn.metrics.f1(y_true, y_hat))
import pickle
from joblib import dump, load
from autosklearn.experimental.askl2 import AutoSklearn2Classifier
automl = AutoSklearn2Classifier(ml_memory_limit=15000, time_left_for_this_task=1800)
automl.fit(X_train, y_train)
y_hat = automl.predict(X_test)

print(automl.cv_results_)
print(automl.sprint_statistics())
print(automl.show_models())
print("Accuracy score ", sklearn.metrics.accuracy_score(y_test, y_hat))
print(automl)
print(automl.cv_results_.keys())
print(automl.cv_results_['param_classifier:__choice__'])
from openpyxl import Workbook, load_workbook

def create_xls(filepath):
    wb = Workbook()
    wb.save(filepath)

def write_xls(filepath, dictionary):
    wb = load_workbook(filepath)
    sheet = wb.active

    headers = list(dictionary.keys()) 

    for index, value in enumerate(headers):
        sheet.cell(row=1, column=index+1).value = value

    for i, x in enumerate(dictionary):
        for idx,value in enumerate(dictionary.values()):
            sheet.cell(row=i+2, column=idx+1).value = value

    wb.save(filepath)
    
create_xls('/kaggle/working/test.xlsx')
write_xls('/kaggle/working/test.xlsx', automl.cv_results_)
cnl_clf = MLPClassifier(solver='adam', alpha=1e-5, hidden_layer_sizes=(10, 5), random_state=randrange(99999), learning_rate = 'invscaling')

cnl_clf.fit(X_train, y_train)

y_pred = cnl_clf.predict(X_test)

print("Accuracy score", sklearn.metrics.accuracy_score(y_test, y_pred))
from autoPyTorch import AutoNetClassification
autoPyTorch = AutoNetClassification("tiny_cs",  # config preset
                                    log_level='info',
                                    max_runtime=300,
                                    min_budget=30,
                                    max_budget=90)

autoPyTorch.fit(X_train, y_train, validation_split=0.3)
y_pred = autoPyTorch.predict(X_test)

print("Accuracy score", accuracy_score(y_test, y_pred))
print('')
print(confusion_matrix(y_true, y_pred))
path  = '/kaggle/input/surface-crack-detection'
data = ImageDataBunch.from_folder(path, train = '.', valid_pct=0.2,
                                  ds_tfms=get_transforms(), size=224,
                                  num_workers=4).normalize(imagenet_stats)

resnet18 = models.resnet18(pretrained=True)
learn = cnn_learner(data, models.resnet18, metrics=[accuracy, f1], model_dir = Path('/kaggle/working'),path = Path("."))

learn.lr_find()
learn.recorder.plot(suggestions=True)

lr1 = 1e-3
lr2 = 1e-1
learn.fit_one_cycle(3,slice(lr1,lr2))
learn.unfreeze()
learn.lr_find()
learn.recorder.plot()
lr = 1e-4
learn.fit_one_cycle(2,lr)
data = []

import os
for dirname, _, filenames in os.walk('/kaggle/input/surface-crack-detection'):
    for filename in filenames:
            data.append(np.asarray(Image.open(os.path.join(dirname, filename))))
            
print(data)
        #print(os.path.join(dirname, filename))
img = open_image('../input/surfacetestimages/Risse_in_der_Wand.jpg')
print(learn.predict(img)[0])
img