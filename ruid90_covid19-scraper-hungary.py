# This Python 3 environment comes with many helpful analytics libraries installed
# It is defined by the kaggle/python Docker image: https://github.com/kaggle/docker-python
# For example, here's several helpful packages to load

import numpy as np # linear algebra
import pandas as pd # data processing, CSV file I/O (e.g. pd.read_csv)

# Input data files are available in the read-only "../input/" directory
# For example, running this (by clicking run or pressing Shift+Enter) will list all files under the input directory

import os
for dirname, _, filenames in os.walk('/kaggle/input'):
    for filename in filenames:
        print(os.path.join(dirname, filename))

# You can write up to 5GB to the current directory (/kaggle/working/) that gets preserved as output when you create a version using "Save & Run All" 
# You can also write temporary files to /kaggle/temp/, but they won't be saved outside of the current session
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
import requests
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np 
from openpyxl import load_workbook
import datetime as dt
#defining the scrapable parts of the site
link = "https://koronavirus.gov.hu/elhunytak"
offset = ['?page=1', '?page=2','?page=3','?page=4','?page=5','?page=6','?page=7','?page=8','?page=9','?page=10','?page=11','?page=12' ,'?page=12',]
result = requests.get (link)
src = result.content
soup = BeautifulSoup(src, 'lxml')
age = soup.find_all("td", {"class": "views-field views-field-field-elhunytak-kor"})
gender = soup.find_all("td", {"class": "views-field views-field-field-elhunytak-nem"})
illness = soup.find_all("td", {"class": "views-field views-field-field-elhunytak-alapbetegsegek"})

#age
for i in range(0,len(offset)):
    result = requests.get (link+offset[i])
    src = result.content
    soup = BeautifulSoup(src, 'lxml')
    age += soup.find_all("td", {"class": "views-field views-field-field-elhunytak-kor"})
ages_string = str(age)
ages_splitted = ages_string.split("<")
del ages_splitted[::2]
ages_splitted_short = [i.replace('td class="views-field views-field-field-elhunytak-kor">','') for i in ages_splitted]
ages_splitted_final= [s.replace(':', '') for s in ages_splitted_short]
agetable = pd.DataFrame(ages_splitted_final)
agetable[0] = agetable[0].astype(int)

#sex
for i in range(0,len(offset)):
    result = requests.get (link+offset[i])
    src = result.content
    soup = BeautifulSoup(src, 'lxml')
    gender += soup.find_all("td", {"class": "views-field views-field-field-elhunytak-nem"})
genders_string = str(gender)
genders_splitted = genders_string.split("<")
del genders_splitted[::2]
genders_splitted_short = [i.replace('td class="views-field views-field-field-elhunytak-nem">','') for i in genders_splitted]
genders_splitted_final= [s.replace(':', '') for s in genders_splitted_short]
gendertable = pd.DataFrame(genders_splitted_final)
gendertable[0] = gendertable[0].astype(str)

#illnesses
for i in range(0,len(offset)):
    result = requests.get (link+offset[i])
    src = result.content
    soup = BeautifulSoup(src, 'lxml')
    illness += soup.find_all("td", {"class": "views-field views-field-field-elhunytak-alapbetegsegek"})
illnesses_string = str(illness)
illnesses_splitted = illnesses_string.split("<")
del illnesses_splitted[::2]
illnesses_splitted_short = [i.replace('td class="views-field views-field-field-elhunytak-alapbetegsegek">','') for i in illnesses_splitted]
illnesses_splitted_final= [s.replace(':', '') for s in illnesses_splitted_short]
illnesstable = pd.DataFrame(illnesses_splitted_final)
illnesstable[0] = illnesstable[0].astype(str)

#creating the concatenated big DF 
bigdf = pd.concat([agetable, gendertable, illnesstable], axis=1, sort=False)
bigdf.columns = ['Age', 'Sex', 'Illness']

#splitting the illnesses column for further EDA  
new = bigdf["Illness"].str.split(",", n = 6, expand = True) 
#removing spaces
bigdf['Sex']=bigdf['Sex'].str.replace(' ','')
for i in range (0,6):
    new[i] = new[i].str.strip()
new.rename(columns={0: "Illness1", 1:"Illness2", 2:"Illness3",3:"Illness4",4:"Illness5",5:"Illness6",6:"Illness7"} ,inplace = True)

#merging dfs
merged = pd.concat([bigdf,new], axis=1, sort=False)
merged.drop(['Illness'], axis = 1, inplace = True)

#calculating the nr. illnesses by removing the nulls 
nulls = merged.isnull().sum(axis=1)
merged['Illnessnumber']= 7 - nulls

#genedrsplit
mergedmale = merged[merged['Sex'].str.contains("Férfi")]
mergedfemale = merged[merged['Sex'].str.contains("Nő")]

##EDA

#searching the most common illnesses (male)
illzmale = mergedmale.iloc[:,2:9].copy()
illzmalevert=illzmale['Illness1'].append(illzmale['Illness2']).reset_index(drop=True)
illzmalevert=illzmalevert.append(illzmale['Illness3']).reset_index(drop=True)
illzmalevert=illzmalevert.append(illzmale['Illness4']).reset_index(drop=True)
illzmalevert=illzmalevert.append(illzmale['Illness5']).reset_index(drop=True)
illzmalevert=illzmalevert.append(illzmale['Illness6']).reset_index(drop=True)
illzmalevert=illzmalevert.append(illzmale['Illness6']).reset_index(drop=True)
illzmalevert.dropna(inplace=True)

#searching the most common illnesses (female)
illzfemale = mergedfemale.iloc[:,2:9].copy()
illzfemalevert=illzfemale['Illness1'].append(illzfemale['Illness2']).reset_index(drop=True)
illzfemalevert=illzfemalevert.append(illzfemale['Illness3']).reset_index(drop=True)
illzfemalevert=illzfemalevert.append(illzfemale['Illness4']).reset_index(drop=True)
illzfemalevert=illzfemalevert.append(illzfemale['Illness5']).reset_index(drop=True)
illzfemalevert=illzfemalevert.append(illzfemale['Illness6']).reset_index(drop=True)
illzfemalevert=illzfemalevert.append(illzfemale['Illness6']).reset_index(drop=True)
illzfemalevert.dropna(inplace=True)

#searching the most common illnesses (total) 
illz = merged.iloc[:,2:9].copy()
illzvert=illz['Illness1'].append(illz['Illness2']).reset_index(drop=True)
illzvert=illzvert.append(illz['Illness3']).reset_index(drop=True)
illzvert=illzvert.append(illz['Illness4']).reset_index(drop=True)
illzvert=illzvert.append(illz['Illness5']).reset_index(drop=True)
illzvert=illzvert.append(illz['Illness6']).reset_index(drop=True)
illzvert=illzvert.append(illz['Illness6']).reset_index(drop=True)

illzvert.dropna(inplace=True)

#top10 illness
top10illz=illzvert.value_counts()[:10].index.tolist()
top10illz=pd.DataFrame(top10illz)

#most common illness
topill=illzvert.value_counts()[:1].index.tolist()

#list os all unique illness
Illzunique=illzvert.unique().tolist()
Illzunique=pd.DataFrame(Illzunique)

#creating new df from the parameters which are calculated 
calculated = pd.DataFrame(columns = ['Date','Total Deaths', 'New Deaths', 'Mean_Age/Total','Mean_Age/Today','Men%', 'Average Number of Illnesses', 'Most Common Illness']) 
calculated.Date = pd.Series(dt.datetime.now())
calculated['Mean_Age/Total'] = merged["Age"].mean()
calculated['Total Deaths'] = merged["Age"].count()
calculated['Most Common Illness']=illzvert.value_counts()[:1].index.tolist()
#share of men in total
a=merged['Sex'].str.count('Férfi').sum()
calculated['Men%'] = a/calculated['Total Deaths']

#avg nr of illnesses
calculated['Average Number of Illnesses'] = merged["Illnessnumber"].mean()
nulls = merged.isnull().sum(axis=1).tolist()

#index adjustment
merged.index = merged.index + 1
calculated.index = calculated.index+1 
with pd.ExcelWriter('covid_death_total_0511.xlsx') as writer:
    merged.to_excel(writer,sheet_name = 'base')
    illzvert.to_excel(writer, sheet_name = 'illztotal')
    top10illz.to_excel(writer, sheet_name = 'top10illz')
    Illzunique.to_excel(writer, sheet_name = 'illzunique')
    illzmalevert.to_excel(writer, sheet_name = 'illzmale')
    illzfemalevert.to_excel(writer, sheet_name = 'illzfemale')
append_df_to_excel('covid_death_calcs.xlsx', calculated, sheet_name='calcs', startrow = None, header = False)
