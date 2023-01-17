import pandas as pd
import schedule
import time
import xlsxwriter 
import openpyxl
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
data_df = pd.read_json('https://finans.truncgil.com/today.json')
data_df = data_df.reindex(["Tür", "Alış", "Satış"])
data_df
data_df.to_excel('sch.xlsx', header=True)
def job():
    book = load_workbook('sch.xlsx')
    writer = pd.ExcelWriter('sch.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    data_df1 = pd.read_json('https://finans.truncgil.com/today.json')
    data_df1 = data_df1.drop('Tür', axis=0)
    for sheetname in writer.sheets:
        data_df1.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, header= False)
    print('Last Run:', datetime.datetime.now())    
    writer.save() 

schedule.every(15).minutes.do(job)
while True:
    schedule.run_pending()
    time.sleep(1)
lastdf = pd.read_excel('sch.xlsx')
lastdf
