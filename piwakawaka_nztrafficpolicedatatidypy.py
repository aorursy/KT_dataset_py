## Offence Data Processor

# this juypter ipython3 notebook is intended to help people process the NZ Police Traffic Offence Data
# published at https://www.police.govt.nz/about-us/publication/road-policing-driver-offence-data
#
# The object is to transform the police workbook so that it conforms with tidy data principles
# https://en.wikipedia.org/wiki/Tidy_data
#
# To do so a new sheet named "Summary" must be added to the police excel workbook
# This program uses the summary data in the workbook to locate the data for transformation
#
# "Summary" sheet layout starting from cell A1
#Table				        Sheet		Headers		Data	Years	Months	Other
#Officer issued red light	Red Light	B4:B42		C4:EH42	B4:B42	C2:EH2	C3:EH3	
#
# where 'Other' refers to speed or alcohol excess values used as index
# Import packages
import pandas as pd
from datetime import datetime
import os
import openpyxl
import re
from openpyxl import load_workbook
from openpyxl import Workbook
# set file name variables
currentfile = 'Policedatato31mar2020.xlsx'
tidydatafile = "tidypolicedataJun20.xlsx"
tidydatatotals = "tidypolicetotalsJun20.xlsx"
p = []         # this is the process log list
p.append("initialised")
p
# Utility Functions
def startrange(excel_range):
    "split the excel range A1:B12 to derive A1"
    splittdr = excel_range.split(sep=':')  # splits the range string at the colon into list
    rangestart = splittdr[0]               # picks first half of the list as string
    return rangestart  # returns a string

def endrange(excel_range):
    "split the excel range A1:B12 to derive B12"
    splittdr = excel_range.split(sep=':')   # splits the range string at the colon into list
    rangeend = splittdr[1]                  # picks first half of the list as string
    return rangeend  # returns a string

def splitcell(s):
    cellist = re.findall('\d+|\D+', s)       # splits the cell into excel column letters and row numbers as list
    return cellist

def breakrange(excel_range):
    a = splitcell(startrange(excel_range))
    b = splitcell(endrange(excel_range))
    c = a + b
    return c

def fatalerror(msg):
    print("FATAL ERROR. processing stopped. Reason:",msg)
    
def status(msg,p):
    p.append(msg)
    return p
# Main processing loop
#
# This function reads the data from the summary sheet in the 
# excel workbook and produces a df which is then saved in various formats
#
def processsheet (tsheet,targetsheetname,hrange,targetdatarange,yrange,mrange,orange):
    "process sheet tidys the tables in each sheet"
    print("processing:"+targetsheetname)
    status("processsheet for"+targetsheetname+" begins",p)
    timeindex = producetimeindex(tsheet,yrange,mrange)
    if orange != None: oindex = produceoindex(tsheet,orange)
    hdr = produceheaders(tsheet,hrange)
    tdf = producerawtdf(tsheet,targetdatarange)
    tdf.columns=hdr
    tdf['timeindex']=timeindex
    if orange != None : tdf['oindex']=oindex
    tdftotals = producetotals(tdf,orange,targetsheetname)
    tdf = tdf[~tdf.timeindex.str.contains("Total")].copy()  # remove the total columns from tdf
    if orange == None : # if there is no other index convert the time index to a datetine
        tdf['timeindex'] = pd.to_datetime(tdf['timeindex'], infer_datetime_format = True)
    else: # otherwise merge the strings in timeindex and oindex to create a single combined index
        tdf['cindex'] = tdf['timeindex']+tdf['oindex']
    tdf['timeindex']
    if orange == None :
        tdf = tdf.set_index(pd.DatetimeIndex(tdf['timeindex']))  
        ## Note the need to tell tdf that it is dealing with date indexes
    else:
        tdf.set_index('cindex')
    tdf.index
    savexl(tdf,tdftotals,orange)
    status("processsheet for"+targetsheetname+" completed",p)
    print("completed")
    # savehdf(tdf,tdftotals,targetsheetname)  - not completed as yet
    # savejson(tdf,tdftotals,targetsheetname) - not completed as yet
    
    
# PART OF PROCESSSHEET

# Function to read the year and month rows and turn them into a continuous
# time series list (including the totals - for the moment)
# we can't convert to a date yet because the total-year rows are still in the data block
#
def producetimeindex(tsheet,yrange,mrange):
    "Extract the time index from the range as a list"
    #mrange and yrange are strings. mrange can be empty e.g police speed excedences
    status("producetimeindex begins",p)
    i=0
    timeindex = []
    yrow = tsheet[startrange(yrange):endrange(yrange)][0] # without index [0] it produces a list of a list of tuples
    if mrange != None:  # i.e mrange is not empty
        mrow = tsheet[startrange(mrange):endrange(mrange)][0] 
        for cell in yrow:
                if yrow[i].value != None: yval = str(yrow[i].value)  # if not empty turn the year cell value into a string = yval
                if mrow[i].value != None: mval = mrow[i].value #note this means the value updated is the last non empty cell
                timeindex.append(mval+"-"+yval)  # create month-year date e.g mar-2010
                i +=1
    else:   #mrange can be empty so there won't be an mrow
            for cell in yrow:
                if yrow[i].value != None: yval = str(yrow[i].value)  # if not empty turn the year cell value into a string = yval
                timeindex.append(yval)  # create month-year date e.g mar-2010
                i +=1
    status("producetimeindex ends",p)
    return timeindex
# PART OF PROCESSSHEET

# Function to copy headers from the excel sheet using the hrange string

def produceheaders(tsheet,hrange):
    "Extract the headers from the sheet as a list"
    status("produce headers started",p)
    hstart = startrange(hrange) # split the headers range
    hend = endrange(hrange)
    hcol = tsheet[hstart:hend] # obtain the excel cells
    ###### derived from data loop below:  
    hdr=[] #  create a list
    i = 0
    for rows in hcol:
        hdr.append(hcol[i][0].value)  # obtain the values from the cells
        i +=1
    status("produceheaders completed",p)
    return hdr
# PART OF PROCESSSHEET

# Function to handle the Other Range

def produceoindex(tsheet,orange):       # before using test if orange = None
    "create the secondary other index as a list"
    status("produce oindex started",p)
    oindex =[]
    orow = tsheet[startrange(orange):endrange(orange)][0]
    i=0  #iterate through the row loading up oindex
    
    for cell in orow:
        if orow[i].value == None: 
            oindex.append("Total") 
        else: oindex.append(orow[i].value)
        i +=1
    status("produce oindex completed",p)
    return oindex
# PART OF PROCESSSHEET

# Function to load up and transpose the data block

def producerawtdf(tsheet,targetdatarange):
    "load up the data block into a data frame"
    status("produce tdf started",p)
    datastart = startrange(targetdatarange)
    dataend = endrange(targetdatarange)
    data_rows = []
    for row in tsheet[datastart:dataend]:
        data_cols = []
        for cell in row:
            if type(cell.value) is str: data_cols.append(int(cell.value))  # if the cell value is a string convert to an integer
            else: data_cols.append(cell.value)                              # else don't
        data_rows.append(data_cols)
    tdf = pd.DataFrame(data_rows)                            # Transform into dataframe
    tdf = tdf.T # transpose to reorient datafrome from horizontal to vertical
    status("produce tdf completed",p)
    return tdf

# PART OF PROCESSSHEET

# Function to derive the totals sheet

def producetotals(tdf,orange,targetsheetname):
    "Produce the totals dataframe out of the tdf frame except when there is an other index (when there is no point)"
    status("produce totals started",p)
    if orange == None: 
        tdftotals = tdf[tdf['timeindex'].str.contains('Total')].copy()   ### note: always make an explicit copy or you are just creating a view
        tdftotals['timeindex'] = tdftotals['timeindex'].apply(lambda s: s.replace('Total-','')) ### apply function to column, lambda is very short function
        tdftotals['timeindex'] = pd.to_datetime(tdftotals['timeindex'], infer_datetime_format = True)
        tdftotals = tdftotals.set_index(pd.DatetimeIndex(tdftotals['timeindex'])) ### and set to index
    else:
        emptytotals = {'col1': [targetsheetname],'col2': ["has no meaningful totals"]}
        tdftotals = pd.DataFrame(emptytotals)
    status("produce totals completed",p)
    return tdftotals
# PART OF PROCESSSHEET

# Function to save data to hdf  {NOT WORKING}

def savehdf(tdf,tdftotals,targetsheetname):
    "Save to HDF for further python processing"
    status("savehdf started",p)
    ttable = targetsheetname[0:29]
    hdfdatafilepath = "hdfdata/"+ttable+"-data.hdf"
    hdftotalsfilepath = "hdftotals/"+ttable+"-total.hdf"
    tdfhdfdata = tdf.to_hdf(path_or_buf=hdfdatafilepath, mode='w', key='tdf', format='table', data_columns='timeindex')
    tdfhdftotal = tdftotals.to_hdf(path_or_buf=hdftotalsfilepath, mode='w', key='tdftotals', format='table', data_columns='timeindex')
    status("save hdf completed",p)
# PART OF PROCESSSHEET

# Function to save data as json file {NOT WORKING}

def savejson(tdf,tdftotals,targetsheetname):
    "Save to JSON for JSON databases"
    status("save json started",p)
    ttable = targetsheetname[0:29]
    jsondatafilepath = "jsondata/"+ttable+"-data.json"
    jsontotalsfilepath = "jsontotals/"+ttable+"-total.json"
    tdfjsondata = tdf.to_json(path_or_buf=jsondatafilepath, orient='columns',date_format='iso',date_unit='s')
    tdfjsontotal = tdftotals.to_json(path_or_buf=jsontotalsfilepath, orient='columns',date_format='iso',date_unit='s')
    status("save json completed",p)
# PART OF PROCESSSHEET

# Function to save data as excel for further processing. {WORKING FINE}

def savexl(tdf,tdftotals,orange):
    "Save to tidy data and total sheets"
    status("save xl started",p)
    ttable = targetsheetname[0:29]
    if orange == None:
        with pd.ExcelWriter(tidydatafile, mode='a', engine='openpyxl',
                     date_format='DD-MMM-YYYY',
                      datetime_format='DD-MMM-YYYY HH:MM:SS') as tdwriter:
            tdf.to_excel(tdwriter, sheet_name=ttable, index=False) 
        with pd.ExcelWriter(tidydatatotals, mode='a', engine='openpyxl',
                     date_format='DD-MMM-YYYY',
                     datetime_format='DD-MMM-YYYY HH:MM:SS') as ttwriter:
             tdftotals.to_excel(ttwriter, sheet_name=ttable, index=False)
    else:
        with pd.ExcelWriter(tidydatafile, mode='a', engine='openpyxl') as tdwriter:
             tdf.to_excel(tdwriter, sheet_name=ttable, index=False) 
        with pd.ExcelWriter(tidydatatotals, mode='a', engine='openpyxl') as ttwriter:
             tdftotals.to_excel(ttwriter, sheet_name=ttable, index=False)
    status("save xl completed",p)
# PART OF PROCESSSHEET

def checksummaryrow(row,passtest):
    tablename = row('Table')
    failmessages = []
    failmessages.append(tablename)
    h = breakrange(row['Headers'])
    d = breakrange(row['Data'])
    y = breakrange(row['Years'])
    m = breakrange(row['Months'])
    if h[1] != d[1]: failmessages.append("header and data start rows dont match")
    if h[3] != d[3]: failmessages.append("header and data end rows dont match")
    if not(d[0]==y[0]==m[o]): failmessages.append("data, years and months dont start with same column")
    if not(d[2]==y[2]==m[2]): failmessages.append("data, years and months dont end with the same column")
    if y[1] != y[3]: failmessages.append("year range not on same row")
    if m[1] != m[3]: failmessages.append("month range not on same row")
    if row['Other'] != None: 
        o = breakrange(row['Other'])
        if not(d[0]==y[0]==m[0]==o[0]): failmessages.append("other start row doesnt match data years or months ")
        if not(d[3]==y[3]==m[3]==o[3]): failmessages.append("other end row doesnt match data years or months ")
        if o[1] != o[3]: failmessages.append("other range not on the same row")
    if len(failmessages)>1: passtest = False
    return failmessages
    
    
# Ensure police data files present and correct
if os.path.isfile(currentfile) == False:
    fatalerror("Specified input police data not found in this path")
    # NEEDS ERROR HANDLER HERE
else:
    policedata = load_workbook(filename = currentfile)
    pds = policedata.sheetnames
    if ("Summary" in pds) == False: 
        fatalerror("You need to prepare a Summary Sheet from the template and add to the police data file")
    # NEEDS ERROR HANDLER HERE
pds
# Prep the Tidy files for receiving the transformed data
#
# Tidy Data Files
#
if os.path.isfile(tidydatafile) == False: #Check to see of the output files exist 
    Tidydata = Workbook()                 #If the output files don't exist create them.
    Tidydata.save(tidydatafile)
    status("tidydata created",p)
else:
    Tidydata = load_workbook(filename = tidydatafile)  # if they do exist flush them of all previous
    for sheet in Tidydata.worksheets:                  # loop through worksheets
        Tidydata.remove(sheet)                         # removing them
    Tidydata.create_sheet() # create one empty sheet
    status("tidydata purged",p)
    Tidydata.save(tidydatafile)
    Tidydata.close()
#
#  Tidy Total files
#
if os.path.isfile(tidydatatotals) == False: #Check to see of the output files exist
    Tidytotals = Workbook()                 #If the output files don't exist create them.
    Tidytotals.save(tidydatatotals)
    status("tidytotals created",p)
else:                                                     #NB if they exist flush previous
    Tidytotals = load_workbook(filename = tidydatatotals)
    for sheet in Tidytotals.worksheets:
        Tidytotals.remove(sheet)
    Tidytotals.create_sheet()
    status("tidytotals purged",p)
    Tidytotals.save(tidydatatotals)
    Tidytotals.close()
# Read the Summary data frame that describes the Police Tables
#
sdf = pd.DataFrame(policedata['Summary'].values)
sdf.columns = sdf.iloc[0]
sdf = sdf.drop(sdf.index[0])
status("summary data frame read and created",p)
sdf
# Defensive parsing of the summary sheet.
# The objective is to identify  input errors  before trying to process them.
# This cell is not complete

passtest = True
reqhdr = {"Table","Sheet","Headers","Data","Years","Months","Other"}
# First ensure the right column has been promoted to column headers and that all the necessary headers are present
sdfhdr = set(sdf.columns.values)
matchhdr = reqhdr.intersection(sdfhdr)
if matchhdr != reqhdr: 
    print("the following necessary headers are missing from the summary:", difference(reqhdr,matchhdr))
    passtest = False  

# Check the data in the Sheet column is the same as the police data sheet names
sdfsheets = ['Summary','Contents']+sdf['Sheet'].unique().tolist()  # add the two worksheets no in the Sheet Column
if pds != sdfsheets:
    pdset = set(pds)
    sdfset = set(sdfsheets)
    difsetpd = pdset - sdfset
    difsetsd = sdfset - pdset
    print("The worksheet names that do not match the names in sheets are", difsetpd)
    print("The names in the summary sheet that do not match the worksheet names are", difsetsd)
    passtest = False

passtest

# MORE Summary Data Sheet TESTS NEEDED
# Now check the values in each row
# for i,row in sdf.iterrows():
# For rows in sdf:
# Match header and data rows
# startrange()
#
# Ensure Year and Month ranges matter

# Passtest not yet conditional on next cell execution
# SET THIS VARIABLE TO RUN THE PRODUCTION LOOP OTHERWISE RUN THE TEST CELL
production = True
#  For loop over all the tables in summary
#
if passtest and production:
    for i,rows in enumerate(sdf.Table): # go through each row in the summary data sheet identifying target tables in main sheets
        tsheet = policedata[sdf.Sheet.iloc[i]] # target police sheet
        targetsheetname = sdf.Table.iloc[i] # string - name of the table on the sheet (there are more than one)
        hrange= sdf.Headers.iloc[i] # string - header row address
        targetdatarange = sdf.Data.iloc[i] # string -Data block
        yrange = sdf.Years.iloc[i] # string - Year range row (mostly empty)
        mrange = sdf.Months.iloc[i] # string - Month range of 12 months plus the unhelpful year end Total
        orange = sdf.Other.iloc[i] # string - Other range (e.g. excess speed or alcohol)
        status(targetsheetname,p)
        processsheet(tsheet,targetsheetname,hrange,targetdatarange,yrange,mrange,orange)  # This does all the work
# TESTBED FOR SINGLE SHEET - run this cell to test data - nb: will not write to xl
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++
if production == False:
    testtable = 1
    tsheet = policedata[sdf.Sheet.iloc[testtable]] # target police sheet
    targetsheetname = sdf.Table.iloc[testtable] # string - name of the table on the sheet (there are more than one)
    hrange= sdf.Headers.iloc[testtable] # string - header row address
    targetdatarange = sdf.Data.iloc[testtable] # string -Data block
    yrange = sdf.Years.iloc[testtable] # string - Year range row (mostly empty)
    mrange = sdf.Months.iloc[testtable] # string - Month range of 12 months plus the unhelpful year end Total
    orange = sdf.Other.iloc[testtable] # string - Other range (e.g. excess speed or alcohol)
    status(targetsheetname,p)
    print("processing:"+targetsheetname)
    status("TEST processsheet for"+targetsheetname+" begins",p)
    timeindex = producetimeindex(tsheet,yrange,mrange)
    if orange != None: oindex = produceoindex(tsheet,orange)
    hdr = produceheaders(tsheet,hrange)
    tdf = producerawtdf(tsheet,targetdatarange)
    tdf.columns=hdr
    tdf['timeindex']=timeindex
    if orange != None : tdf['oindex']=oindex
    tdftotals = producetotals(tdf,orange,targetsheetname)
    tdf = tdf[~tdf.timeindex.str.contains("Total")].copy()  # remove the total columns from tdf
    if orange == None : # if there is no other index convert the time index to a datetine
        tdf['timeindex'] = pd.to_datetime(tdf['timeindex'], infer_datetime_format = True)
    else: # otherwise merge the strings in timeindex and oindex to create a single combined index
        tdf['cindex'] = tdf['timeindex']+tdf['oindex']
    tdf['timeindex']
    if orange == None :
        tdf = tdf.set_index(pd.DatetimeIndex(tdf['timeindex']))  
            ## Note the need to tell tdf that it is dealing with date indexes
    else:
        tdf.set_index('cindex')
    status(targetsheetname +" TEST PROCESS Completed ",p)
    tdf.index
    tdf
print(p)



