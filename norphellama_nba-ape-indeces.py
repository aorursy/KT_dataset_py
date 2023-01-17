%matplotlib inline
from fastai.basics import *
import numpy as np
import pandas as pd
import openpyxl
from pathlib import Path
xlsx_file = Path('../input/nba-players-measurements-19472017/Player - Bio Stats (1947-2017).xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file) 

# Read the active sheet:
sheet = wb_obj.active
x = torch.ones(sheet.max_row,2);
y = torch.ones(sheet.max_row,1);
ht_min = 150;
ws_min = 160;
rowsFilled = 0;
for row in range(2, sheet.max_row + 1):
    ht_val = sheet["H" + str(row)].value;
    ws_val = sheet["I" + str(row)].value;
    if(ht_val != None and ht_val > ht_min):
        if(ws_val != None and  ws_val > ws_min):
            x[rowsFilled, 0] = float(ht_val); 
            y[rowsFilled] = float(ws_val);
            rowsFilled = rowsFilled + 1;
x = x[0:rowsFilled,:];
y = y[0:rowsFilled];
plt.scatter(x[:,0], y);
def mse(y_hat, y): return ((y_hat-y)**2).mean()
a = tensor(-1.,1);
y_hat = x@a
mse(y_hat, y)
plt.scatter(x[:,0],y)
plt.scatter(x[:,0],y_hat);
a = tensor(-1.,1);
a = nn.Parameter(a); a
def update():
    y_hat = x@a
    loss = mse(y, y_hat)
    if t % 10 == 0: print(loss)
    loss.backward()
    with torch.no_grad():
        a.sub_(lr * a.grad)
        a.grad.zero_()
lr = 1e-6
for t in range(100): update();
plt.scatter(x[:,0],y)
with torch.no_grad():
    plt.scatter(x[:,0],x@a);