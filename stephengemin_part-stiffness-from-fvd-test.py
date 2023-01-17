import pandas as pd

import numpy as np

import scipy.signal

import openpyxl

from openpyxl.chart import Reference, Series

from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font

import os, time

import matplotlib.pyplot as plt



import os

print(os.listdir("../input"))
def check_dir_exists(dir_path):

    try:

        os.mkdir(dir_path)

    except OSError:

        pass

        #dir already exists

    return dir_path



def read_results_sheet(infile, cols):

    temp_df = pd.read_excel(infile, sheet='Results')

    temp_df = temp_df[cols]

    return list(temp_df[cols[0]]), temp_df



def write_data_to_single_worksheet(writer, sheet, col_counter):

    print (sheet)

    temp_df = pd.read_excel(in_file,sheet_name = sheet,usecols="A:C")

    temp_bool = temp_df['Load'] >= 5000

    rows = temp_bool.idxmax()

    temp_df = temp_df.drop(range(rows-1))

    temp_df['Position'] = temp_df['Position'] - temp_df.iloc[0]['Position']

    temp_df['Corrected Position'] = temp_df['Corrected Position'] - temp_df.iloc[0]['Corrected Position']

    temp_df['Corrected Position_smooth'] = scipy.signal.savgol_filter(temp_df['Corrected Position'], window_length = 101, polyorder=1)

    temp_df.to_excel(writer,sheet_name = "Data", index=False, startrow = 1, startcol=col_counter, freeze_panes=(2,0))

    

    return temp_df, len(temp_df.columns)



def write_unique_key_header(out_file, sheetnames):

    cell_col = 1

    wb = openpyxl.load_workbook(out_file)

    ws = wb['Data']

    for sheet in sheetnames:

        ws.cell(row=1, column=cell_col).value = sheet

        cell_col += len(temp_df.columns)

        wb.save(out_file)

    

def get_stiffness(x, y):

    par = np.polyfit(x, y, 1, full=True)

    slope = par[0][0] 

    yint = par[0][1]

    

    #R² value

    variance = np.var(y)

    residuals = np.var([(slope*xx + yint - yy) for xx,yy in zip(x,y)])

    r_sqr = np.round(1-residuals/variance, decimals=2)    

    return slope, yint, r_sqr
def create_overview_plot(temp_df, FEA_df, stats_summary, output_file):

    counter=0

    fig, ax = plt.subplots()

    fig.set_size_inches(10,6)

    for j in range(0,temp_df.shape[1],4):

        ax.plot(temp_df.iloc[:,j+3], temp_df.iloc[:,j], 'blue', linewidth=1, label=stats_summary.loc[counter, 'Specimen Identification'] + '_Corrected')

        ax.plot(temp_df.iloc[:,j+1], temp_df.iloc[:,j], 'green', linewidth=1, label=stats_summary.loc[counter, 'Specimen Identification'])

        counter += 1

    

    # Add rotor and load cell FEA results

    ax.plot(FEA_df.loc[:,'Part Displacement'], FEA_df.loc[:,'Part Force'], 'r-', linewidth=1, label='Part FEA')

    ax.plot(FEA_df.loc[:,'LCell Displacement'], FEA_df.loc[:,'LCell Force'], '-', color='orange', linewidth=1, label='Load Cell FEA')    

    h, l = ax.get_legend_handles_labels()

    ax.legend(h, l, loc='lower right', shadow=True)

    

    ax.set_ylabel('Force [N]', fontsize=12)

    ax.tick_params(axis='y', labelsize=12)

    ax.set_ylim(0,50000)

    ax.set_xlabel('Position [mm]', fontsize=12)

    ax.set_xlim(0,0.35)

    ax.set_title('Force vs. Corrected Position', fontsize=16)

    ax.grid(which='major', linestyle='-')

    ax.minorticks_on()

    ax.grid(which='minor', linestyle=':')



    plt.tight_layout(pad=1.5)

    plt.savefig(user_desktop_temp + 'Summary.pdf')



def create_overview_plot_inexcel(out_file):

    wb = openpyxl.load_workbook(out_file)

    ws = wb['Data']

    ws2 = wb['FEA']

    ws3 = wb['Summary']

    tempo = pd.read_excel(writer, sheet_name='Data', skiprows=1)

    stats = pd.read_excel(writer, sheet_name='Summary')



    # tempo.shape

    c1 = openpyxl.chart.ScatterChart()



    c1.style = 2

    c1.legend.position = 't'

    c1.width = 20

    c1.height = 11

    c1.title='Force vs Position Summary'

    c1.x_axis.title = 'Position [mm]'

    c1.x_axis.scaling.min = 0

    c1.x_axis.scaling.max = 0.15

    c1.y_axis.title = 'Force [N]'

    c1.y_axis.scaling.min = 0

    c1.y_axis.scaling.max = 50000



    col_counter = 1

    for items in stats['UniqueStamp']:

        xvalues = Reference(ws, min_col=col_counter+3, min_row=3, max_row=tempo.shape[0])

        yvalues = Reference(ws, min_col=col_counter, min_row=3, max_row=tempo.shape[0])

        series = Series(yvalues, xvalues, title=items)

        series.smooth = True

        c1.append(series)

        col_counter += 4



    #FEA rotor

    xvalues = Reference(ws2, min_col=2, min_row=2, max_row=1000)

    yvalues = Reference(ws2, min_col=1, min_row=2, max_row=1000)

    series = Series(yvalues, xvalues, title='Rotor FEA')

    series.smooth = True

    c1.append(series)



    #FEA rotor

    xvalues = Reference(ws2, min_col=4, min_row=2, max_row=1000)

    yvalues = Reference(ws2, min_col=3, min_row=2, max_row=1000)

    series = Series(yvalues, xvalues, title='Load Cell FEA')

    series.smooth = True

    c1.append(series)



    ws3.add_chart(c1, "J2")

    wb.save(out_file)
#I normally use this coding on my computer, to output any results to a temp folder on my Desktop

# Get location for the Desktop 

# so this will work on any Windows computer

user_desktop = os.path.expanduser("~/Desktop/").replace("\\","/")

user_desktop_temp = check_dir_exists(user_desktop + "Temp/")



#there will be an error from the file paths below.  In a computer everything is in the same relative directory.  

# Kaggle works a little different, so it doesn't work within this environment.

in_file = "Tests Compiled 4.xlsx"

FEA_file = "FEA results.xlsm"

out_file = user_desktop_temp + "Output.xlsx"

cols = [ 'UniqueStamp', 'Test Number', 'Specimen Identification', 'StartDate']

col_counter = 0



writer = pd.ExcelWriter(out_file)



add_cols = ['Slope [kN/mm]','Y-int [kN]','r_sqr']

#loop through all tests to create "Data" worksheet

sheetnames, stats_summary = read_results_sheet(in_file, cols)

for col in add_cols:

    stats_summary[col] = None

    

#loop through all tests to create "Summary" worksheet

for i in range(len(sheetnames)):

    temp_df, shift = write_data_to_single_worksheet(writer, sheetnames[i], col_counter)

    #define range to analyze the stiffness

    min_load = 10000 #N

    max_load = 40000 #N

    lin_start_row = (temp_df['Load']>min_load).idxmax()

    lin_end_row = (temp_df['Load']>max_load).idxmax()

    

    # get line best fit

    x = np.array(temp_df.loc[lin_start_row:lin_end_row, 'Corrected Position'])

    y = np.array(temp_df.loc[lin_start_row:lin_end_row, 'Load'])

    stats = list(get_stiffness(x,y))

    stats[0] /= 1000

    stats[1] /= 1000

    stats_summary.loc[i, add_cols] = stats

    col_counter += shift

stats_summary.to_excel(writer, sheet_name='Summary', index=False)



#create "FEA" worksheet summary

FEA_df = pd.read_excel(FEA_file)

stats = list(get_stiffness(np.array(FEA_df['Part Displacement']), np.array(FEA_df['Part Force'])))

FEA_df.loc[0,'Part Stiffness [kN/mm]'] = stats[0] / 1000

stats = list(get_stiffness(np.array(FEA_df['LCell Displacement']), np.array(FEA_df['LCell Force'])))

FEA_df.loc[0,'Load Cell Stiffness [kN/mm]'] = stats[0] / 1000

FEA_df.to_excel(writer, sheet_name='FEA', index=False)

writer.save()



#Add UniqueStamp to the "Data" worksheet

write_unique_key_header(out_file, sheetnames)



#create overview plot

temp_df = pd.read_excel(out_file, sheet_name='Data', skiprows=1)

create_overview_plot(temp_df, FEA_df, stats_summary, user_desktop_temp + 'Summary.pdf')

create_overview_plot_inexcel(out_file)



#create worksheet to transfer data to FEA request

to_drop = pd.concat([temp_df[temp_df.columns[1::4]],temp_df[temp_df.columns[2::4]]], axis=1)

temp_df.drop(axis=1, columns = to_drop.columns, inplace=True)

temp_df.to_excel(writer, sheet_name='FvD for FEA', index=False, startrow=2, freeze_panes=(2,0))

writer.save()